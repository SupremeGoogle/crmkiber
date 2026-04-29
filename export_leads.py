#!/usr/bin/env python3
import csv
import html
import json
import re
import ssl
import sys
import urllib.parse
import urllib.request
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
CONFIG = {
    "base_url": "https://kiberonekaliningrad.s20.online",
    "company_id": 1,
    # ОБНОВЛЯЙ при истечении сессии:
    "cookie": "supportOnlineTalkID=08798423424f5bd02976401edc252a2c; PHPSESSID=g5rl7p448vaputca6delp27h65; _csrf=93ef67846c9a5a5d282c7fa9a323ff0572e8add11c9e45313acad20df9788f7aa%3A2%3A%7Bi%3A0%3Bs%3A5%3A%22_csrf%22%3Bi%3A1%3Bs%3A32%3A%222WMTtIcA2N2VQmJS2Vs2ygsDtKL_Fleo%22%3B%7D",
    "csrf_header": "v8Nn4DTx25F8e6gwmWwnhNEnWc22ZzUgLHOZbymq4_-NlCq0QLi40E41mmbIAW3X43Eq_88ARmRYONUwb8aGkA==",
    "csrf_form": "Cgs0Gx1PpbcUKg2c5hrnJgysGIMp-Hs2fqQWMk3CBIE4XHlPaQbG9iZkP8q3d611PvprsVCfCHIK71ptC65h7g==",
    # Период отчета
    "date_from": "01.12.2022",
    "date_to": "30.04.2026",
}
SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE
def lead_score(obj):
    if not isinstance(obj, dict):
        return 0
    score = 0
    if "id" in obj:
        score += 2
    if any(k in obj for k in ("name", "clientName", "firstName", "lastName")):
        score += 2
    if any(k in obj for k in ("email", "clientEmail", "emails")):
        score += 1
    if any(k in obj for k in ("phone", "clientPhone", "phones")):
        score += 1
    if any(k in obj for k in ("status", "statusName", "leadStatus")):
        score += 1
    if any(k in obj for k in ("createdAt", "updatedAt", "dateAdd", "date")):
        score += 1
    return score
def pick_items(data):
    queue = [data]
    best = []
    best_score = -1
    seen = set()
    while queue:
        node = queue.pop(0)
        node_id = id(node)
        if node_id in seen:
            continue
        seen.add(node_id)
        if isinstance(node, list):
            sample = node[:40]
            avg = (sum(lead_score(x) for x in sample) / len(sample)) if sample else 0
            rank = avg * 1000 + len(node)
            if rank > best_score:
                best_score = rank
                best = node
            for item in sample:
                if isinstance(item, (dict, list)):
                    queue.append(item)
        elif isinstance(node, dict):
            for val in node.values():
                if isinstance(val, (dict, list)):
                    queue.append(val)
    return best if isinstance(best, list) else []


def strip_tags(text):
    no_tags = re.sub(r"<[^>]+>", " ", text or "")
    return html.unescape(re.sub(r"\s+", " ", no_tags)).strip()


def normalize_phone_text(raw):
    if not raw:
        return ""
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 11 and digits[0] in ("7", "8"):
        return f"+7 ({digits[1:4]}) {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
    if len(digits) == 10 and digits[0] == "9":
        return f"+7 ({digits[0:3]}) {digits[3:6]}-{digits[6:8]}-{digits[8:10]}"
    return ""


def extract_phone(row_html, joined_text):
    patterns = [
        joined_text,
        " ".join(re.findall(r"href=[\"']tel:([^\"']+)[\"']", row_html, flags=re.IGNORECASE)),
        " ".join(
            re.findall(
                r"(?:data-phone|data-value|value)=[\"']([^\"']*)[\"']",
                row_html,
                flags=re.IGNORECASE,
            )
        ),
    ]

    for src in patterns:
        if not src:
            continue
        for candidate in re.findall(r"\+?\d[\d\s()\-]{8,}\d", src):
            phone = normalize_phone_text(candidate)
            if phone:
                return phone
    return ""


def parse_leads_from_report_html(content_html):
    rows = re.findall(r"<tr[^>]*>([\s\S]*?)</tr>", content_html or "", flags=re.IGNORECASE)
    leads = []

    for row in rows:
        cells = re.findall(r"<td[^>]*>([\s\S]*?)</td>", row, flags=re.IGNORECASE)
        if len(cells) < 3:
            continue

        cell_text = [strip_tags(c) for c in cells]
        if not any(cell_text):
            continue

        # Header-like rows skip
        joined = " ".join(cell_text).lower()
        if "дата добавления" in joined and "наименование" in joined:
            continue

        id_match = re.search(r"/lead/view/(\d+)", row, flags=re.IGNORECASE)
        lead_id = int(id_match.group(1)) if id_match else None
        if lead_id is None:
            first_num = re.search(r"\b(\d{2,})\b", cell_text[0] if cell_text else "")
            lead_id = int(first_num.group(1)) if first_num else None

        # Typical columns: # | date | name | stage | ...
        date_value = cell_text[1] if len(cell_text) > 1 else ""
        name_value = cell_text[2] if len(cell_text) > 2 else ""
        status_value = cell_text[3] if len(cell_text) > 3 else ""

        joined_text = " ".join(cell_text)
        email_match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", joined_text)

        phone = extract_phone(row, joined_text)

        if not (lead_id or name_value or email_match or phone):
            continue

        leads.append(
            {
                "id": lead_id,
                "name": name_value or (f"Лид #{lead_id}" if lead_id else "Без имени"),
                "email": email_match.group(0) if email_match else "",
                "phone": phone,
                "status": status_value,
                "source": "report/lead-created",
                "date": date_value,
            }
        )

    return leads
def post_report():
    back_url = urllib.parse.quote(
        f"{CONFIG['base_url']}/company/{CONFIG['company_id']}/lead/index?LeadSearch%5Bf_removed%5D=2",
        safe="",
    )
    url = f"{CONFIG['base_url']}/company/{CONFIG['company_id']}/report/lead-created?backUrl={back_url}"
    form = urllib.parse.urlencode(
        {
            "_csrf": CONFIG["csrf_form"],
            "ReportForm[d1]": CONFIG["date_from"],
            "ReportForm[d2]": CONFIG["date_to"],
            "ReportForm[pipelines]": "",
            "ReportForm[sort]": "1",
            "ReportForm[is_skip_archive]": "0",
            "export": "",
        }
    ).encode("utf-8")
    req = urllib.request.Request(url, data=form, method="POST")
    req.add_header("Accept", "*/*")
    req.add_header("Origin", CONFIG["base_url"])
    req.add_header("Referer", f"{CONFIG['base_url']}/company/{CONFIG['company_id']}/lead/index?LeadSearch%5Bf_removed%5D=2")
    req.add_header("Cookie", CONFIG["cookie"])
    req.add_header("X-Requested-With", "XMLHttpRequest")
    req.add_header("X-CSRF-Token", CONFIG["csrf_header"])
    req.add_header("Content-Type", "application/x-www-form-urlencoded; charset=UTF-8")
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=90) as resp:
        body = resp.read().decode("utf-8", errors="ignore")
        ctype = resp.headers.get("Content-Type", "")
    return body, ctype
def get_field(lead, names):
    for name in names:
        val = lead.get(name)
        if val:
            return val
    return None
def normalize_phone(value):
    if not value:
        return None
    digits = re.sub(r"\D", "", str(value))
    return digits[-10:] if len(digits) >= 7 else None
def normalize_email(value):
    return str(value).strip().lower() if value else None
def dedup(leads):
    sorted_leads = sorted(
        leads,
        key=lambda x: str(get_field(x, ("createdAt", "dateAdd", "date", "updatedAt")) or ""),
        reverse=True,
    )
    seen_phone = {}
    seen_email = {}
    result = []
    removed = 0
    total = len(sorted_leads)
    for i, lead in enumerate(sorted_leads, start=1):
        email = normalize_email(get_field(lead, ("email", "clientEmail")))
        phone = normalize_phone(get_field(lead, ("phone", "clientPhone")))
        lead_id = lead.get("id", "?")
        lead_name = get_field(lead, ("name", "clientName")) or "Без имени"
        print(f"[{i}/{total}] Лид ID={lead_id} | {lead_name}", flush=True)
        is_dup = (phone and phone in seen_phone) or (email and email in seen_email)
        if is_dup:
            removed += 1
            print("  -> дубль, пропущен", flush=True)
            continue
        if phone:
            seen_phone[phone] = True
        if email:
            seen_email[email] = True
        result.append(lead)
        print("  -> добавлен", flush=True)
    return result, removed
def to_csv_rows(leads):
    rows = []
    for lead in leads:
        name = get_field(lead, ("name", "clientName"))
        if not name and (lead.get("firstName") or lead.get("lastName")):
            name = f"{lead.get('firstName', '')} {lead.get('lastName', '')}".strip()
        rows.append(
            {
                "ID": lead.get("id", ""),
                "Имя": name or f"Лид #{lead.get('id', '')}",
                "Email": get_field(lead, ("email", "clientEmail")) or "",
                "Телефон": get_field(lead, ("phone", "clientPhone")) or "",
                "Статус": get_field(lead, ("statusName", "status", "leadStatus")) or "",
                "Источник": get_field(lead, ("sourceName", "source", "referrer")) or "",
                "Дата": get_field(lead, ("createdAt", "dateAdd", "date", "updatedAt")) or "",
            }
        )
    return rows


def save_xlsx(rows, xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Лиды"

    headers = ["ID", "Имя", "Email", "Телефон", "Статус", "Источник", "Дата"]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    col_limits = {
        1: (8, 14),
        2: (24, 48),
        3: (22, 42),
        4: (18, 24),
        5: (16, 26),
        6: (16, 28),
        7: (14, 18),
    }

    for col in range(1, ws.max_column + 1):
        max_len = len(str(ws.cell(row=1, column=col).value or ""))
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            txt = str(val) if val is not None else ""
            max_len = max(max_len, len(txt))
        min_w, max_w = col_limits.get(col, (12, 40))
        ws.column_dimensions[get_column_letter(col)].width = max(min_w, min(max_w, max_len + 2))

    wb.save(xlsx_path)
def main():
    print("Загружаем архивных лидов из CRM...")
    body, ctype = post_report()
    try:
        parsed = json.loads(body)
    except json.JSONDecodeError:
        print("Ошибка: CRM вернула не JSON. Скорее всего устарели cookie/CSRF.")
        sys.exit(1)
    leads = pick_items(parsed)
    if not leads and isinstance(parsed, dict) and isinstance(parsed.get("content"), str):
        leads = parse_leads_from_report_html(parsed["content"])
    print(f"Получено лидов из источника: {len(leads)}")
    if isinstance(parsed, dict):
        print(f"Ключи ответа: {list(parsed.keys())[:30]}")
    if not leads:
        print("Лидов 0. Обнови cookie + csrf_header + csrf_form из Network.")
        with open("debug_response_when_zero.json", "w", encoding="utf-8") as f:
            json.dump(parsed, f, ensure_ascii=False, indent=2)
        print("Сохранил ответ CRM: debug_response_when_zero.json")
        sys.exit(1)
    unique_leads, removed = dedup(leads)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = f"leads_raw_{ts}.json"
    csv_path = f"leads_unique_{ts}.csv"
    xlsx_path = f"leads_unique_{ts}.xlsx"
    print("Сохраняю JSON...")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(leads, f, ensure_ascii=False, indent=2)
    rows = to_csv_rows(unique_leads)
    print("Сохраняю XLSX...")
    save_xlsx(rows, xlsx_path)
    print("Сохраняю CSV...")
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f, fieldnames=["ID", "Имя", "Email", "Телефон", "Статус", "Источник", "Дата"]
        )
        writer.writeheader()
        writer.writerows(rows)
    print("")
    print(f"Всего лидов: {len(leads)}")
    print(f"Уникальных: {len(unique_leads)}")
    print(f"Удалено дублей: {removed}")
    print(f"RAW JSON: {json_path}")
    print(f"XLSX: {xlsx_path}")
    print(f"CSV: {csv_path}")
if __name__ == "__main__":
    main()
