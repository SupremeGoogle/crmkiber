#!/usr/bin/env python3
import csv
import html
import json
import re
import ssl
import sys
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


APP_DIR = Path(__file__).resolve().parent
SESSION_FILE = APP_DIR / "crm_session.json"

CONFIG = {
    "base_url": "https://kiberonekaliningrad.s20.online",
    "company_id": 1,
    "date_from": "01.12.2022",
    "date_to": "30.04.2026",
}


SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE


def load_session():
    if SESSION_FILE.exists():
        try:
            return json.loads(SESSION_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_session(session):
    SESSION_FILE.write_text(json.dumps(session, ensure_ascii=False, indent=2), encoding="utf-8")


def ask_value(title, current):
    print(f"\n{title}")
    if current:
        print("(Enter = оставить текущее значение)")
    value = input("> ").strip()
    return value or current


def strip_tags(text):
    no_tags = re.sub(r"<[^>]+>", " ", text or "")
    return html.unescape(re.sub(r"\s+", " ", no_tags)).strip()


def normalize_phone_text(raw):
    if not raw:
        return ""
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 11 and digits[0] in ("7", "8"):
        return f"+7 ({digits[1:4]}) {digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
    if len(digits) == 10 and digits[0] == "9":
        return f"+7 ({digits[0:3]}) {digits[3:6]}-{digits[6:8]}-{digits[8:10]}"
    return ""


def extract_phone(row_html, joined_text):
    patterns = [
        joined_text,
        " ".join(re.findall(r"href=[\"']tel:([^\"']+)[\"']", row_html, flags=re.IGNORECASE)),
        " ".join(
            re.findall(
                r"(?:data-phone|data-value|value)=[\"']([^\"']*)[\"']",
                row_html,
                flags=re.IGNORECASE,
            )
        ),
    ]

    for src in patterns:
        if not src:
            continue
        for candidate in re.findall(r"\+?\d[\d\s()\-]{8,}\d", src):
            phone = normalize_phone_text(candidate)
            if phone:
                return phone
    return ""


def fetch_phone_from_lead_card(lead_id, session):
    if not lead_id:
        return ""
    url = f"{CONFIG['base_url']}/company/{CONFIG['company_id']}/lead/view/{lead_id}"
    req = urllib.request.Request(url, method="GET")
    req.add_header("Accept", "text/html,application/xhtml+xml")
    req.add_header("Referer", f"{CONFIG['base_url']}/company/{CONFIG['company_id']}/lead/index?LeadSearch%5Bf_removed%5D=2")
    req.add_header("Cookie", session["cookie"])
    req.add_header("User-Agent", "Mozilla/5.0")

    try:
        with urllib.request.urlopen(req, context=SSL_CTX, timeout=45) as resp:
            html_body = resp.read().decode("utf-8", errors="ignore")
    except Exception:
        return ""

    patterns = [
        r"Мобильный\s*</[^>]*>\s*<[^>]*>\s*([^<]+)",
        r"Мобильный\s*[:\-]?\s*([^<\n\r]+)",
        r"phone\"\s*:\s*\"([^\"]+)\"",
    ]
    for p in patterns:
        m = re.search(p, html_body, flags=re.IGNORECASE)
        if not m:
            continue
        normalized = normalize_phone_text(m.group(1))
        if normalized:
            return normalized

    around_mobile = re.search(r"Мобильный[\s\S]{0,200}", html_body, flags=re.IGNORECASE)
    if around_mobile:
        m2 = re.search(r"\+?\d[\d\s()\-]{8,}\d", around_mobile.group(0))
        if m2:
            normalized = normalize_phone_text(m2.group(0))
            if normalized:
                return normalized

    return ""


def enrich_missing_phones(leads, session):
    missing = [lead for lead in leads if not (lead.get("phone") or "").strip() and lead.get("id")]
    if not missing:
        return 0

    print(f"\nДогружаем телефоны из карточек лидов: {len(missing)} шт.")
    filled = 0
    total = len(missing)
    for idx, lead in enumerate(missing, start=1):
        lead_id = lead.get("id")
        print(f"[{idx}/{total}] Проверяю ID={lead_id}...", flush=True)
        phone = fetch_phone_from_lead_card(lead_id, session)
        if phone:
            lead["phone"] = phone
            filled += 1
            print(f"  -> найден телефон: {phone}", flush=True)
        else:
            print("  -> телефон не найден", flush=True)

    print(f"Дозаполнено телефонов: {filled}\n")
    return filled


def parse_leads_from_report_html(content_html):
    rows = re.findall(r"<tr[^>]*>([\s\S]*?)</tr>", content_html or "", flags=re.IGNORECASE)
    leads = []

    for row in rows:
        cells = re.findall(r"<td[^>]*>([\s\S]*?)</td>", row, flags=re.IGNORECASE)
        if len(cells) < 3:
            continue

        cell_text = [strip_tags(c) for c in cells]
        joined_text = " ".join(cell_text)

        id_match = re.search(r"/lead/view/(\d+)", row, flags=re.IGNORECASE)
        lead_id = int(id_match.group(1)) if id_match else None

        date_value = cell_text[1] if len(cell_text) > 1 else ""
        name_value = cell_text[2] if len(cell_text) > 2 else ""
        status_value = cell_text[3] if len(cell_text) > 3 else ""

        email_match = re.search(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", joined_text)

        phone = extract_phone(row, joined_text)

        if not (lead_id or name_value or email_match or phone):
            continue

        leads.append(
            {
                "id": lead_id,
                "name": name_value or (f"Лид #{lead_id}" if lead_id else "Без имени"),
                "email": email_match.group(0) if email_match else "",
                "phone": phone,
                "status": status_value,
                "source": "report/lead-created",
                "date": date_value,
            }
        )

    return leads


def post_report(session):
    back_url = urllib.parse.quote(
        f"{CONFIG['base_url']}/company/{CONFIG['company_id']}/lead/index?LeadSearch%5Bf_removed%5D=2",
        safe="",
    )
    url = f"{CONFIG['base_url']}/company/{CONFIG['company_id']}/report/lead-created?backUrl={back_url}"

    form = urllib.parse.urlencode(
        {
            "_csrf": session["csrf_form"],
            "ReportForm[d1]": CONFIG["date_from"],
            "ReportForm[d2]": CONFIG["date_to"],
            "ReportForm[pipelines]": "",
            "ReportForm[sort]": "1",
            "ReportForm[is_skip_archive]": "0",
            "export": "",
        }
    ).encode("utf-8")

    req = urllib.request.Request(url, data=form, method="POST")
    req.add_header("Accept", "*/*")
    req.add_header("Origin", CONFIG["base_url"])
    req.add_header("Referer", f"{CONFIG['base_url']}/company/{CONFIG['company_id']}/lead/index?LeadSearch%5Bf_removed%5D=2")
    req.add_header("Cookie", session["cookie"])
    req.add_header("X-Requested-With", "XMLHttpRequest")
    req.add_header("X-CSRF-Token", session["csrf_header"])
    req.add_header("Content-Type", "application/x-www-form-urlencoded; charset=UTF-8")

    with urllib.request.urlopen(req, context=SSL_CTX, timeout=90) as resp:
        body = resp.read().decode("utf-8", errors="ignore")
    return json.loads(body)


def get_field(lead, names):
    for name in names:
        val = lead.get(name)
        if val:
            return val
    return None


def normalize_phone(value):
    if not value:
        return None
    digits = re.sub(r"\D", "", str(value))
    return digits[-10:] if len(digits) >= 7 else None


def normalize_email(value):
    return str(value).strip().lower() if value else None


def dedup(leads):
    sorted_leads = sorted(leads, key=lambda x: str(get_field(x, ("date", "createdAt", "updatedAt")) or ""), reverse=True)
    seen_phone = set()
    seen_email = set()
    result = []
    removed = 0

    total = len(sorted_leads)
    for i, lead in enumerate(sorted_leads, start=1):
        email = normalize_email(get_field(lead, ("email", "clientEmail")))
        phone = normalize_phone(get_field(lead, ("phone", "clientPhone")))
        print(f"[{i}/{total}] ID={lead.get('id', '?')} | {get_field(lead, ('name', 'clientName')) or 'Без имени'}", flush=True)

        if (phone and phone in seen_phone) or (email and email in seen_email):
            removed += 1
            print("  -> дубль, пропущен", flush=True)
            continue

        if phone:
            seen_phone.add(phone)
        if email:
            seen_email.add(email)
        result.append(lead)
        print("  -> добавлен", flush=True)

    return result, removed


def to_csv_rows(leads):
    rows = []
    for lead in leads:
        rows.append(
            {
                "ID": lead.get("id", ""),
                "Имя": get_field(lead, ("name", "clientName")) or "",
                "Email": get_field(lead, ("email", "clientEmail")) or "",
                "Телефон": get_field(lead, ("phone", "clientPhone")) or "",
                "Статус": get_field(lead, ("statusName", "status", "leadStatus")) or "",
                "Источник": get_field(lead, ("sourceName", "source", "referrer", "source")) or "",
                "Дата": get_field(lead, ("date", "createdAt", "dateAdd", "updatedAt")) or "",
            }
        )
    return rows


def save_xlsx(rows, xlsx_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Лиды"

    headers = ["ID", "Имя", "Email", "Телефон", "Статус", "Источник", "Дата"]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Авто-ширина столбцов с разумными лимитами
    min_max_width = {
        1: (8, 14),   # ID
        2: (24, 48),  # Имя
        3: (22, 42),  # Email
        4: (18, 24),  # Телефон
        5: (16, 26),  # Статус
        6: (16, 28),  # Источник
        7: (14, 18),  # Дата
    }

    for col in range(1, ws.max_column + 1):
        max_len = len(str(ws.cell(row=1, column=col).value or ""))
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            txt = str(val) if val is not None else ""
            max_len = max(max_len, len(txt))
        min_w, max_w = min_max_width.get(col, (12, 40))
        width = max(min_w, min(max_w, max_len + 2))
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(xlsx_path)


def main():
    print("=== CRM Export EXE ===")
    session = load_session()
    session["cookie"] = ask_value("Вставь Cookie из Network (строка целиком):", session.get("cookie", ""))
    session["csrf_header"] = ask_value("Вставь X-CSRF-Token:", session.get("csrf_header", ""))
    session["csrf_form"] = ask_value("Вставь _csrf из form-data:", session.get("csrf_form", ""))
    save_session(session)

    print("\nЗагружаем лидов...")
    parsed = post_report(session)
    leads = parse_leads_from_report_html(parsed.get("content", ""))
    print(f"Получено лидов: {len(leads)}")

    if not leads:
        print("Лидов 0. Скорее всего, токены устарели. Вставь свежие значения из Network.")
        sys.exit(1)

    enrich_missing_phones(leads, session)

    unique_leads, removed = dedup(leads)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = APP_DIR / f"leads_unique_{ts}.csv"
    xlsx_path = APP_DIR / f"leads_unique_{ts}.xlsx"
    rows = to_csv_rows(unique_leads)

    save_xlsx(rows, xlsx_path)

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["ID", "Имя", "Email", "Телефон", "Статус", "Источник", "Дата"])
        writer.writeheader()
        writer.writerows(rows)

    print("\nГотово")
    print(f"Всего лидов: {len(leads)}")
    print(f"Уникальных: {len(unique_leads)}")
    print(f"Удалено дублей: {removed}")
    print(f"XLSX: {xlsx_path}")
    print(f"CSV:  {csv_path}")
    input("\nНажми Enter для выхода...")


if __name__ == "__main__":
    main()
